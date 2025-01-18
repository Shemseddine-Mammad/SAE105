import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.styles import Font, PatternFill, Alignment

class AnalyseExcel:
    def __init__(self, fichier_csv):
        """Initialisation avec le fichier CSV des séances"""
        self.fichier_csv = fichier_csv
        self.wb = Workbook()
        self.ws_seances = self.wb.active
        self.ws_seances.title = "Séances"
        self.ws_stats = self.wb.create_sheet("Statistiques")

    def charger_et_formater_donnees(self):
        """Charge les données CSV et les formate dans Excel"""
        try:
            # Lecture du CSV avec pandas
            df = pd.read_csv(self.fichier_csv, delimiter=';')
            
            # Formatage de la feuille des séances
            self.ws_seances.append(['Date', 'Durée', 'Type'])
            for index, row in df.iterrows():
                self.ws_seances.append([row['Date'], row['Durée'], row['Type']])
            
            # Mise en forme
            for cell in self.ws_seances[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Ajustement des colonnes
            for column in self.ws_seances.columns:
                max_length = 0
                for cell in column:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                self.ws_seances.column_dimensions[column[0].column_letter].width = max_length + 2

            return True
        except Exception as e:
            print(f"Erreur lors du chargement des données : {str(e)}")
            return False

    def calculer_statistiques(self):
        """Calcule et ajoute les statistiques"""
        try:
            # Comptage par type de séance
            self.ws_stats.append(['Statistiques par type de séance'])
            self.ws_stats.append(['Type', 'Nombre de séances'])
            
            types_count = {}
            for row in self.ws_seances.iter_rows(min_row=2):
                type_seance = row[2].value
                types_count[type_seance] = types_count.get(type_seance, 0) + 1
            
            for type_seance, count in types_count.items():
                self.ws_stats.append([type_seance, count])
            
            # Mise en forme
            self.ws_stats['A1'].font = Font(bold=True, size=12)
            for cell in self.ws_stats[2]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            return True
        except Exception as e:
            print(f"Erreur lors du calcul des statistiques : {str(e)}")
            return False

    def ajouter_graphiques(self):
        """Ajoute des graphiques pour visualiser les données"""
        try:
            # Graphique en barres
            chart_bar = BarChart()
            chart_bar.title = "Répartition des séances par type"
            chart_bar.style = 10
            
            data = Reference(self.ws_stats, min_row=3, max_row=self.ws_stats.max_row,
                           min_col=2, max_col=2)
            cats = Reference(self.ws_stats, min_row=3, max_row=self.ws_stats.max_row,
                           min_col=1, max_col=1)
            
            chart_bar.add_data(data)
            chart_bar.set_categories(cats)
            
            self.ws_stats.add_chart(chart_bar, "E2")
            
            # Graphique en camembert
            chart_pie = PieChart()
            chart_pie.title = "Distribution des types de séances"
            chart_pie.style = 10
            
            chart_pie.add_data(data)
            chart_pie.set_categories(cats)
            
            self.ws_stats.add_chart(chart_pie, "E18")
            
            return True
        except Exception as e:
            print(f"Erreur lors de l'ajout des graphiques : {str(e)}")
            return False

    def sauvegarder(self, nom_fichier):
        """Sauvegarde le classeur Excel"""
        try:
            self.wb.save(nom_fichier)
            return True
        except Exception as e:
            print(f"Erreur lors de la sauvegarde : {str(e)}")
            return False

def main():
    # Fichier CSV généré par automation.py
    fichier_csv = "test_dcpdump.csv"  # À modifier avec le bon chemin
    fichier_excel = "analyse_r107.xlsx"
    
    # Création de l'analyse Excel
    analyse = AnalyseExcel(fichier_csv)
    
    if analyse.charger_et_formater_donnees():
        print("Données chargées et formatées avec succès")
        
        if analyse.calculer_statistiques():
            print("Statistiques calculées avec succès")
            
            if analyse.ajouter_graphiques():
                print("Graphiques ajoutés avec succès")
                
                if analyse.sauvegarder(fichier_excel):
                    print(f"Analyse Excel sauvegardée dans : {fichier_excel}")
                else:
                    print("Erreur lors de la sauvegarde du fichier Excel")
            else:
                print("Erreur lors de l'ajout des graphiques")
        else:
            print("Erreur lors du calcul des statistiques")
    else:
        print("Erreur lors du chargement des données")

if __name__ == "__main__":
    main()